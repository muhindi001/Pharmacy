from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ExcelReportGenerator:

    @staticmethod
    def generate(title, headers, rows):

        workbook = Workbook()

        worksheet = workbook.active
        worksheet.title = title

        # Report Title
        worksheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=len(headers),
        )

        cell = worksheet.cell(row=1, column=1)

        cell.value = title
        cell.font = Font(size=16, bold=True)
        cell.alignment = Alignment(horizontal="center")

        # Header Style
        header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )

        header_font = Font(
            bold=True,
            color="FFFFFF",
        )

        row_number = 3

        for index, header in enumerate(headers, start=1):

            cell = worksheet.cell(
                row=row_number,
                column=index,
            )

            cell.value = header
            cell.fill = header_fill
            cell.font = header_font

        # Data
        row_number += 1

        for row in rows:

            for column, value in enumerate(row, start=1):

                worksheet.cell(
                    row=row_number,
                    column=column,
                ).value = value

            row_number += 1

        # Auto Width
        for column in worksheet.columns:

            length = 0

            column_letter = get_column_letter(
                column[0].column
            )

            for cell in column:

                try:

                    if len(str(cell.value)) > length:

                        length = len(str(cell.value))

                except Exception:
                    pass

            worksheet.column_dimensions[
                column_letter
            ].width = length + 5

        output = BytesIO()

        workbook.save(output)

        output.seek(0)

        return output.getvalue()
    